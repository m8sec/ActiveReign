from openpyxl import Workbook, load_workbook

from ar3.pysmb.file_ops import get_fileobj, close_fileobj
from ar3.ops.enum.file_parser.parse_regex import regex_search


def parse_xlsx(xlsx_headers, regex, max_size, max_chars, timeout, con, share, path, filename):
    # Check column headers against XLSX_PARSER & regex first 20 lines of file
    rcount = 1
    try:
        file_obj = get_fileobj(con, share, path, filename, max_size, timeout)
        wb = load_workbook(filename=file_obj, read_only=True)
        for ws in wb.sheetnames:
            ws = wb[ws]
            for row in ws.rows:
                print(row)
                # Due to the format of xlsx docs on parser first 10 lines
                if rcount == 10: break # Move to next sheet
                for cell in row:
                    if cell.value != None:
                        cell = cell.value.rstrip().lstrip()
                        # Perform header lookup
                        if cell.lower() in xlsx_headers:
                            close_fileobj(file_obj)
                            return { 'Parser': 'Excel',
                                     'ParserDetails': 'Keyword',
                                     'LineCount': str(rcount),
                                     'LineSample': str(cell).strip('\n')}
                        else:
                             # Perform regex search on data
                            search_data = regex_search(regex, max_chars, cell, rcount, filename)
                            if search_data:
                                close_fileobj(file_obj)
                                return search_data
                # Next Row
                rcount += 1
        close_fileobj(file_obj)
    except:
        pass
    return False